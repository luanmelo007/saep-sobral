"""
scripts/migrar_planilha.py  — VERSÃO CORRIGIDA
Migração completa da planilha SAEP Sobral para o Supabase.

Uso:
    python3 scripts/migrar_planilha.py --arquivo "scripts/Plano de Ação SAEP - SOBRAL....xlsx" --limpar

Variáveis de ambiente:
    SUPABASE_URL, SUPABASE_KEY

Antes de usar --limpar, rode UMA VEZ no SQL Editor do Supabase as políticas
de DELETE (veja bloco no final deste arquivo).
"""

import os
import sys
import argparse
import re
import warnings
from datetime import datetime, date

import openpyxl
from supabase import create_client

warnings.simplefilter("ignore")

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

# ── Helpers ───────────────────────────────────────────────────────────────────

def normalizar_data(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
            try:
                return datetime.strptime(valor.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def normalizar_codigo(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        try:
            return str(int(valor))
        except (ValueError, OverflowError):
            return None
    s = str(valor).strip()
    s = re.split(r"[\s\-_\n]", s)[0]
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    try:
        return str(int(digits))
    except ValueError:
        return None


def normalizar_status(valor, opcoes):
    if not valor:
        return opcoes[0]
    v = str(valor).strip().upper()
    mapa = {
        "REALIZADO": "REALIZADO", "OK": "REALIZADO",
        "EM ANDAMENTO": "EM ANDAMENTO", "ANDAMENTO": "EM ANDAMENTO",
        "PLANEJADO": "PLANEJADO", "ATRASADO": "ATRASADO",
        "NÃO REALIZADO": "NÃO REALIZADO", "NAO REALIZADO": "NÃO REALIZADO",
        "CANCELADO": "CANCELADO",
    }
    for k, mapped in mapa.items():
        if k in v and mapped in opcoes:
            return mapped
    return opcoes[0]


def normalizar_escala(valor):
    if valor is None:
        return "NÃO APLICADA"
    v = str(valor).strip().upper()
    mapa = {
        "AV": "AV", "AVANÇADO": "AV", "AD": "AD", "ADEQUADO": "AD",
        "B": "B", "BÁSICO": "B", "BASICO": "B",
        "AB": "AB", "ABAIXO DO BÁSICO": "AB", "ABAIXO": "AB",
        "NÃO APLICADA": "NÃO APLICADA", "NAO APLICADA": "NÃO APLICADA",
        "#N/A": "NÃO APLICADA", "NÃO REALIZOU": "NÃO APLICADA", "-%": "NÃO APLICADA",
    }
    return mapa.get(v, "NÃO APLICADA")


def texto(valor, maxlen=2000):
    if valor is None:
        return None
    s = str(valor).strip()
    return s[:maxlen] if s and s not in ("None", "nan") else None


def lotes(lst, n=50):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


MENTOR_KEYWORDS = {
    "luan": "Luan", "raimundo": "Raimundo", "lucilani": "Lucilani",
    "décio": "Décio", "decio": "Décio", "claudecir": "Claudecir",
}

def resolver_mentor(valor, mentor_map):
    if not valor:
        return None
    v = str(valor).lower()
    for kw, nome in MENTOR_KEYWORDS.items():
        if kw in v and nome in mentor_map:
            return mentor_map[nome]
    return None


CURSO_KEYWORDS = {
    "eletrotécnica": "ELETROTÉCNICA", "eletrotecnica": "ELETROTÉCNICA",
    "eletroeletronica": "ELETROTÉCNICA",
    "eletromecânica": "ELETROMECÂNICA", "eletromecanica": "ELETROMECÂNICA", "eletromec": "ELETROMECÂNICA",
    "mecatrônica": "MECATRÔNICA", "mecatronica": "MECATRÔNICA", "mecatron": "MECATRÔNICA",
    "desenvolvimento de sistemas": "DS", " ds": "DS",
    "redes": "REDES",
    "planejamento e controle": "PCP", "pcp": "PCP",
    "segurança": "TST", "seguranca": "TST", "sst": "TST", "tst": "TST",
    "mecânica": "MECÂNICA", "mecanica": "MECÂNICA",
}

def resolver_curso(valor, curso_map):
    if not valor:
        return None
    v = " " + str(valor).lower()
    for kw, codigo in CURSO_KEYWORDS.items():
        if kw in v and codigo in curso_map:
            return curso_map[codigo]
    return None


ABA_PARA_CURSO = {
    "DS": "DS", "Eletromecânica": "ELETROMECÂNICA", "Eletrotécnica": "ELETROTÉCNICA",
    "Mecânica": "MECÂNICA", "Mecatrônica": "MECATRÔNICA", "PCP": "PCP",
    "REDES": "REDES", "TST": "TST",
}


def limpar_transacionais(sb):
    for tabela in ("resultados_alunos", "historico_acoes", "avaliacoes", "acoes"):
        try:
            sb.table(tabela).delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        except Exception as e:
            print(f"  ⚠️  Não foi possível limpar {tabela}: {e}")
    print("  ✅ Tabelas transacionais limpas")


def buscar_mentores(sb):
    res = sb.table("mentores").select("id, nome").execute()
    return {r["nome"]: r["id"] for r in res.data}


CURSOS_DADOS = [
    {"codigo": "ELETROTÉCNICA",  "nome_completo": "TÉCNICO EM ELETROTÉCNICA",                      "modalidade": "PRESENCIAL"},
    {"codigo": "ELETROMECÂNICA", "nome_completo": "TÉCNICO EM ELETROMECÂNICA",                     "modalidade": "PRESENCIAL"},
    {"codigo": "MECATRÔNICA",    "nome_completo": "TÉCNICO EM MECATRÔNICA",                        "modalidade": "PRESENCIAL"},
    {"codigo": "MECÂNICA",       "nome_completo": "TÉCNICO EM MECÂNICA",                           "modalidade": "PRESENCIAL"},
    {"codigo": "DS",             "nome_completo": "TÉCNICO EM DESENVOLVIMENTO DE SISTEMAS",        "modalidade": "PRESENCIAL"},
    {"codigo": "REDES",          "nome_completo": "TÉCNICO EM REDES DE COMPUTADORES",              "modalidade": "PRESENCIAL"},
    {"codigo": "PCP",            "nome_completo": "TÉCNICO EM PLANEJAMENTO E CONTROLE DA PRODUÇÃO","modalidade": "PRESENCIAL"},
    {"codigo": "TST",            "nome_completo": "TÉCNICO EM SEGURANÇA DO TRABALHO",              "modalidade": "PRESENCIAL"},
]

def importar_cursos(sb):
    res = sb.table("cursos").select("id, codigo").execute()
    existentes = {r["codigo"]: r["id"] for r in res.data}
    if existentes:
        print(f"  ℹ️  {len(existentes)} cursos já existem")
        return existentes
    sb.table("cursos").insert(CURSOS_DADOS).execute()
    res = sb.table("cursos").select("id, codigo").execute()
    mapa = {r["codigo"]: r["id"] for r in res.data}
    print(f"  ✅ {len(mapa)} cursos inseridos")
    return mapa


def importar_turmas(sb, wb, curso_map, mentor_map):
    ws = wb["Turmas Elegíveis"]
    rows = list(ws.iter_rows(values_only=True))
    vistos = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        codigo = normalizar_codigo(row[0])
        if not codigo:
            continue
        nome_curso = texto(row[1], 100) or ""
        up = nome_curso.upper()
        turno = ("EAD" if "EAD" in up else "NEM" if "NEM" in up else
                 "NOITE" if "NOITE" in up else "TARDE" if "TARDE" in up else "MANHÃ")
        vistos[codigo] = {
            "codigo": codigo,
            "curso_id": resolver_curso(nome_curso, curso_map),
            "turno": turno,
            "data_inicio": normalizar_data(row[2]),
            "data_fim": normalizar_data(row[3]),
            "ciclo_saep": texto(row[5], 20) if row[5] and "NÃO" not in str(row[5]).upper() else None,
            "status_saep": normalizar_status(str(row[6] or ""), ["PLANEJADO", "REALIZADO", "NÃO ELEGÍVEL"]),
        }
    unicos = list(vistos.values())
    for batch in lotes(unicos):
        sb.table("turmas").upsert(batch, on_conflict="codigo").execute()
    res = sb.table("turmas").select("id, codigo").execute()
    mapa = {r["codigo"]: r["id"] for r in res.data}
    print(f"  ✅ {len(unicos)} turmas inseridas ({len(mapa)} total)")
    return mapa


def garantir_turma(sb, codigo, curso_id, turma_map):
    if codigo in turma_map:
        return turma_map[codigo]
    res = sb.table("turmas").upsert(
        {"codigo": codigo, "curso_id": curso_id, "turno": "MANHÃ", "status_saep": "NÃO ELEGÍVEL"},
        on_conflict="codigo").execute()
    novo_id = res.data[0]["id"] if res.data else None
    turma_map[codigo] = novo_id
    return novo_id


def importar_acoes_macro(sb, wb, mentor_map, curso_map):
    ws = wb["PA InterlocutorMentor"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next((i for i, r in enumerate(rows) if r[0] == "CURSO" or r[2] == "O QUE"), 1)
    registros = []
    for row in rows[header_idx + 1:]:
        o_que = texto(row[2])
        if not o_que:
            continue
        status_raw = texto(row[1], 30) or texto(row[11], 30) or "PLANEJADO"
        registros.append({
            "tipo": "MACRO", "o_que": o_que, "porque": texto(row[3]),
            "onde": texto(row[4], 200), "quem_id": resolver_mentor(str(row[5] or ""), mentor_map),
            "quando_previsto": normalizar_data(row[6]), "como": texto(row[7]),
            "status": normalizar_status(status_raw, ["PLANEJADO", "EM ANDAMENTO", "REALIZADO", "ATRASADO", "NÃO REALIZADO"]),
            "data_realizacao": normalizar_data(row[8]), "observacoes": texto(row[9]),
            "encaminhamentos": texto(row[10]), "curso_id": resolver_curso(str(row[0] or ""), curso_map),
        })
    for batch in lotes(registros):
        sb.table("acoes").insert(batch).execute()
    print(f"  ✅ {len(registros)} ações macro inseridas")


def importar_acoes_turmas(sb, wb, mentor_map, curso_map, turma_map):
    ws = wb["Plano de Ação - Turmas"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next((i for i, r in enumerate(rows) if r[0] == "CÓD TURMA"), 0)
    registros = []
    for row in rows[header_idx + 1:]:
        o_que = texto(row[3])
        if not o_que:
            continue
        primeiro_cod = normalizar_codigo(str(row[0] or "").split("\n")[0])
        turma_id = turma_map.get(primeiro_cod)
        registros.append({
            "tipo": "TURMA" if turma_id else "MACRO", "turma_id": turma_id,
            "curso_id": resolver_curso(str(row[1] or ""), curso_map), "o_que": o_que,
            "porque": texto(row[4]), "onde": texto(row[5], 200),
            "quem_id": resolver_mentor(str(row[6] or ""), mentor_map),
            "quando_previsto": normalizar_data(row[7]), "como": texto(row[8]),
            "status": normalizar_status(str(row[2] or ""), ["PLANEJADO", "EM ANDAMENTO", "REALIZADO", "ATRASADO", "NÃO REALIZADO"]),
            "data_realizacao": normalizar_data(row[9]), "observacoes": texto(row[10]),
        })
    for batch in lotes(registros):
        sb.table("acoes").insert(batch).execute()
    print(f"  ✅ {len(registros)} ações de turmas inseridas")


def importar_avaliacoes(sb, wb, mentor_map, turma_map):
    ws = wb["Avaliações"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next((i for i, r in enumerate(rows) if r[1] == "CÓD TURMA"), 1)
    registros = []
    for row in rows[header_idx + 1:]:
        numero = texto(row[3], 50)
        if not numero:
            continue
        cod = normalizar_codigo(row[1])
        turma_id = turma_map.get(cod)
        if not turma_id:
            continue
        tipo = ("PRATICA_OFICIAL" if numero in ("PO", "PP") else
                "SIMULADO" if "simulado" in numero.lower() else "OBJETIVA")
        resultado = None
        if row[8] and str(row[8]).strip() not in ("", "None", "-"):
            try:
                v = float(str(row[8]).replace("%", "").replace(",", ".").strip())
                resultado = v if v <= 1.0 else v / 100.0
            except (ValueError, TypeError):
                pass
        registros.append({
            "turma_id": turma_id, "numero": numero, "tipo": tipo,
            "local": texto(row[4], 200), "responsavel_id": resolver_mentor(str(row[5] or ""), mentor_map),
            "data_prevista": normalizar_data(row[6]), "ciclo": normalizar_data(row[0]),
            "status": normalizar_status(str(row[7] or ""), ["PLANEJADO", "REALIZADO", "ATRASADO", "CANCELADO"]),
            "resultado_medio": resultado, "observacoes": texto(row[9]),
        })
    for batch in lotes(registros):
        sb.table("avaliacoes").insert(batch).execute()
    print(f"  ✅ {len(registros)} avaliações inseridas")


def garantir_avaliacao(sb, turma_id, numero, aval_cache):
    chave = (turma_id, numero)
    if chave in aval_cache:
        return aval_cache[chave]
    res = sb.table("avaliacoes").insert({
        "turma_id": turma_id, "numero": numero, "tipo": "OBJETIVA", "status": "REALIZADO",
    }).execute()
    novo_id = res.data[0]["id"] if res.data else None
    aval_cache[chave] = novo_id
    return novo_id


def importar_resultados(sb, wb, curso_map, turma_map):
    res = sb.table("avaliacoes").select("id, turma_id, numero").execute()
    aval_cache = {(r["turma_id"], r["numero"]): r["id"] for r in res.data}
    total = 0
    for aba, curso_codigo in ABA_PARA_CURSO.items():
        if aba not in wb.sheetnames:
            continue
        curso_id = curso_map.get(curso_codigo)
        ws = wb[aba]
        rows = list(ws.iter_rows(values_only=True))
        turma_atual = None
        col_aluno = None
        mapa_colunas = []
        registros = []
        vistos = set()
        for row in rows:
            achou = False
            for cell in row:
                m = re.search(r"\d{5}\.\d{4}\.\d{4}", str(cell or ""))
                if m:
                    cod = normalizar_codigo(m.group(0))
                    turma_atual = garantir_turma(sb, cod, curso_id, turma_map)
                    col_aluno = None
                    mapa_colunas = []
                    achou = True
                    break
            if achou:
                continue
            idx_aluno = next((i for i, c in enumerate(row) if "Aluno(a)" in str(c or "")), None)
            if idx_aluno is not None:
                col_aluno = idx_aluno
                mapa_colunas = []
                for i, cell in enumerate(row):
                    s = str(cell or "").strip()
                    if "diagnóstica" in s.lower() or "simulado" in s.lower() or s in ("PO", "PP"):
                        num = re.sub(r"\d{1,2}/\d{1,2}/\d{2,4}", "", s)
                        num = re.sub(r"\(.*?\)", "", num).replace("\n", " ").strip()
                        mapa_colunas.append((i, num))
                continue
            if turma_atual is None or col_aluno is None or not mapa_colunas:
                continue
            nome = texto(row[col_aluno], 200) if col_aluno < len(row) else None
            if not nome or "média" in nome.lower() or "aluno(a)" in nome.lower():
                continue
            for idx_pont, numero in mapa_colunas:
                if idx_pont >= len(row):
                    continue
                esc = normalizar_escala(row[idx_pont + 1] if idx_pont + 1 < len(row) else None)
                if esc == "NÃO APLICADA":
                    continue
                pontuacao = None
                if row[idx_pont] is not None:
                    try:
                        v = float(str(row[idx_pont]).replace(",", ".").replace("%", ""))
                        pontuacao = v * 1000 if 0 < v <= 1 else v
                    except (ValueError, TypeError):
                        pass
                aval_id = garantir_avaliacao(sb, turma_atual, numero, aval_cache)
                if not aval_id:
                    continue
                chave = (aval_id, nome)
                if chave in vistos:
                    continue
                vistos.add(chave)
                registros.append({
                    "avaliacao_id": aval_id, "nome_aluno": nome,
                    "pontuacao": pontuacao, "escala": esc,
                })
        if registros:
            for batch in lotes(registros):
                try:
                    sb.table("resultados_alunos").upsert(batch, on_conflict="avaliacao_id,nome_aluno").execute()
                except Exception as e:
                    print(f"  ⚠️  Erro no lote de {aba}: {e}")
            total += len(registros)
            print(f"    • {aba}: {len(registros)} resultados")
    print(f"  ✅ {total} resultados de alunos inseridos")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--arquivo", required=True)
    parser.add_argument("--limpar", action="store_true")
    args = parser.parse_args()
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("❌ Configure SUPABASE_URL e SUPABASE_KEY")
        sys.exit(1)
    print("🔄 Conectando ao Supabase...")
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    print(f"📖 Abrindo planilha: {args.arquivo}")
    wb = openpyxl.load_workbook(args.arquivo, read_only=True, data_only=True)
    print("\n👥 Buscando mentores...")
    mentor_map = buscar_mentores(sb)
    print(f"  ✅ {len(mentor_map)} mentores: {list(mentor_map.keys())}")
    if args.limpar:
        print("\n🧹 Limpando tabelas transacionais...")
        limpar_transacionais(sb)
    print("\n📚 Importando cursos...")
    curso_map = importar_cursos(sb)
    print("\n🏫 Importando turmas...")
    turma_map = importar_turmas(sb, wb, curso_map, mentor_map)
    print("\n📋 Importando ações macro...")
    importar_acoes_macro(sb, wb, mentor_map, curso_map)
    print("\n📋 Importando ações por turma...")
    importar_acoes_turmas(sb, wb, mentor_map, curso_map, turma_map)
    print("\n📊 Importando avaliações...")
    importar_avaliacoes(sb, wb, mentor_map, turma_map)
    print("\n👤 Importando resultados por aluno...")
    importar_resultados(sb, wb, curso_map, turma_map)
    print("\n🎉 Migração concluída!")


if __name__ == "__main__":
    main()

# ═══════════════════════════════════════════════════════════════════════════════
# SQL necessário no Supabase (rode UMA VEZ no SQL Editor para usar --limpar):
#
# CREATE POLICY "delete_acoes"      ON acoes              FOR DELETE USING (true);
# CREATE POLICY "delete_avaliacoes" ON avaliacoes         FOR DELETE USING (true);
# CREATE POLICY "delete_resultados" ON resultados_alunos  FOR DELETE USING (true);
# CREATE POLICY "delete_historico"  ON historico_acoes    FOR DELETE USING (true);
# ═══════════════════════════════════════════════════════════════════════════════